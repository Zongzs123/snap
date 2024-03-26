# This is a sample Python script.
import time
import datetime
import os
import glob
import sys
import math
from random import shuffle, random, randint

import tensorflow as tf
from keras.layers import *
from keras.models import *
from keras.optimizers import *
from keras.callbacks import *
from keras.wrappers.scikit_learn import KerasRegressor
from keras import backend as K
import numpy as np

from sklearn.model_selection import KFold
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import ShuffleSplit, cross_val_score
from sklearn.model_selection import train_test_split
from sklearn.metrics import make_scorer
from keras.utils import to_categorical
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

# preprocess of the data
h, l, t, dd, ddd = 2.5, 2.0, 0.8, 1.0, 1.0

columns = ["L", "w", "w^3", "alpha", "num_vert", "num_hori", "s1", "s2", "s3", "s4", "s5", "s6", "s7", "s8", "s9",
           "s10", "s11", "s12", "s13", "s14", "s15", "s16", "s17", "s18", "s19", "s20", "s21", "s22", "s23", "s24",
           "s25", "s26", "s27", "s28", "s29", "s30", "s31"]
filename = 'processed-data-of-1.6-2.2.xlsx'
data1_1 = pd.read_excel(filename, sheet_name=1, skiprows=None, names=columns)
data1_2 = pd.read_excel(filename, sheet_name=2, skiprows=None, names=columns)
data1_3 = pd.read_excel(filename, sheet_name=3, skiprows=None, names=columns)
data1_4 = pd.read_excel(filename, sheet_name=4, skiprows=None, names=columns)
data1_5 = pd.read_excel(filename, sheet_name=5, skiprows=None, names=columns)
data1 = pd.concat([data1_1, data1_2, data1_3, data1_4, data1_5], axis=0)
data1.pop('s1')
data2 = data1.to_numpy(dtype=np.float32)
data_X = data2[:, :4]
data_Y = data2[:, 6:]

scaler1 = StandardScaler()
scaler1.fit(data_X)
data_X_standarized = scaler1.transform(data_X)

scaler2 = StandardScaler()
scaler2.fit(data_Y)
data_Y_standarized = scaler2.transform(data_Y)

tmp_array = []
for i in range(data2.shape[0]):
    tmp = data_X_standarized[i, :].tolist()
    tmp += to_categorical(data2[i, 4] - 1, 4).tolist()
    tmp += to_categorical(data2[i, 5] - 1, 4).tolist()
    tmp_array += [tmp]
data_X_one_hot = np.array(tmp_array, dtype=np.float32)
data3 = np.append(data_X_one_hot, data_Y_standarized, axis=-1)
data3_augmented = data3

nn = 5
for i in range(1654):
    for j in range(4):
        for k in range(4):
            for m in range(nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1 - tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1 - (tmp1 + tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1 - (tmp1 + tmp2 + tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
# 第二种波形
for i in range(1654, 1845):
    for j in range(4):
        for k in range(4):
            for m in range(9 * nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1 - tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1 - (tmp1 + tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1 - (tmp1 + tmp2 + tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
# 第三种波形
for i in range(1845, 1967):
    for j in range(4):
        for k in range(4):
            for m in range(14 * nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1 - tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1 - (tmp1 + tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1 - (tmp1 + tmp2 + tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
# 第四种波形
for i in range(1967, 2054):
    for j in range(4):
        for k in range(4):
            for m in range(19 * nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1 - tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1 - (tmp1 + tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1 - (tmp1 + tmp2 + tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]

for i in range(2054, 2104):
    for j in range(4):
        for k in range(4):
            for m in range(33 * nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1 - tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1 - (tmp1 + tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1 - (tmp1 + tmp2 + tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]

print(data3_augmented.shape)

np.random.shuffle(data3_augmented)
inputs = data3_augmented[:, :12]
targets = data3_augmented[:, 12:42]


# save augmented data to designated file
# def createwb(wbname):
#     wb=openpyxl.Workbook()
#     wb.save(filename=wbname)
#
#
# def data_toExcel(data, wbname):
#     wb = openpyxl.load_workbook(wbname)
#     worksheet1 = wb.create_sheet(index=0, title='sheet1')
#     for i in range(len(data)):
#         for j in range(len(data[i, :])):
#             worksheet1.cell(row=i+1, column=j+1).value = data[i, j]
#     wb.save(wbname)
#     wb.close()
#
#
# augmented_data_file = 'l2-augmented-1108.xlsx'
# createwb(augmented_data_file)
# data_toExcel(data3_augmented, augmented_data_file)


# Define loss function
def loss_function(y_true, y_pred):
    return K.mean(K.square(y_true - y_pred))


def loss_MAE(y_true, y_pred):
    return K.mean(K.abs(y_true - y_pred))


# Train forward regressor
# def forward_model():
#     inp = Input(shape=(11,), name='surrogator_input')
#
#     x = Dense(128, activation='relu')(inp)
#     x = BatchNormalization()(x)
#
#     x = Dense(256, activation='relu')(x)
#     x = BatchNormalization()(x)
#
#     out = Dense(16, activation=None)(x)
#
#     model = Model(inputs=inp, outputs=out, name='surrogator_points_model')
#
#     return model

# Function to create the forward network
def forward_network():
    inp = Input(shape=(12,), name='forward_input')
    x = inp
    x = Dense(300)(x)
    x = BatchNormalization()(x)
    x = Activation('relu')(x)
    x = Dense(1024)(x)
    x = BatchNormalization()(x)
    x = Activation('relu')(x)
    out = Dense(30, activation=None)(x)
    model = Model(inputs=inp, outputs=out)

    model.compile(loss=loss_function, optimizer=Adam(learning_rate=0.001), metrics=[loss_MAE])

    return model


# Split your data into train and test sets
inputs_train, inputs_test, targets_train, targets_test = train_test_split(inputs, targets, test_size=0.2,
                                                                          random_state=42)

# Define a learning rate scheduler callback
lr_scheduler = ReduceLROnPlateau(monitor='loss', factor=0.1, patience=10, min_lr=0.000001, verbose=1)

print('------------------------------------------------------------------------')
print(f'Training ...')

kfold = KFold(n_splits=5, shuffle=True, random_state=42)
count = 0

# Create array to save history data
loss_MAE_size16 = []
val_loss_MAE_size16 = []
loss_MAE_size32 = []
val_loss_MAE_size32 = []
loss_MAE_size64 = []
val_loss_MAE_size64 = []
loss_MAE_size128 = []
val_loss_MAE_size128 = []
loss_MAE_size256 = []
val_loss_MAE_size256 = []

for train_index, val_index in kfold.split(inputs_train):
    train_inputs, train_targets = inputs_train[train_index, :], targets_train[train_index, :]
    val_inputs, val_targets = inputs_train[val_index, :], targets_train[val_index, :]
    # Create the forward model
    forward_model_size16 = forward_network()
    forward_model_size32 = forward_network()
    forward_model_size64 = forward_network()
    forward_model_size128 = forward_network()
    forward_model_size256 = forward_network()

    # Train each model using the training data
    results_size16 = forward_model_size16.fit(train_inputs, train_targets,
                                              batch_size=16,
                                              epochs=200,
                                              callbacks=[lr_scheduler],
                                              verbose=2,
                                              validation_data=(val_inputs, val_targets),
                                              shuffle=True)
    results_size32 = forward_model_size32.fit(train_inputs, train_targets,
                                              batch_size=32,
                                              epochs=200,
                                              callbacks=[lr_scheduler],
                                              verbose=2,
                                              validation_data=(val_inputs, val_targets),
                                              shuffle=True)
    results_size64 = forward_model_size64.fit(train_inputs, train_targets,
                                              batch_size=64,
                                              epochs=200,
                                              callbacks=[lr_scheduler],
                                              verbose=2,
                                              validation_data=(val_inputs, val_targets),
                                              shuffle=True)
    results_size128 = forward_model_size128.fit(train_inputs, train_targets,
                                                batch_size=128,
                                                epochs=200,
                                                callbacks=[lr_scheduler],
                                                verbose=2,
                                                validation_data=(val_inputs, val_targets),
                                                shuffle=True)
    results_size256 = forward_model_size256.fit(train_inputs, train_targets,
                                                batch_size=256,
                                                epochs=200,
                                                callbacks=[lr_scheduler],
                                                verbose=2,
                                                validation_data=(val_inputs, val_targets),
                                                shuffle=True)

    loss_MAE_size16.append(results_size16.history['loss_MAE'])
    val_loss_MAE_size16.append(results_size16.history['val_loss_MAE'])
    loss_MAE_size32.append(results_size32.history['loss_MAE'])
    val_loss_MAE_size32.append(results_size32.history['val_loss_MAE'])
    loss_MAE_size64.append(results_size64.history['loss_MAE'])
    val_loss_MAE_size64.append(results_size64.history['val_loss_MAE'])
    loss_MAE_size128.append(results_size128.history['loss_MAE'])
    val_loss_MAE_size128.append(results_size128.history['val_loss_MAE'])
    loss_MAE_size256.append(results_size256.history['loss_MAE'])
    val_loss_MAE_size256.append(results_size256.history['val_loss_MAE'])

    count += 1

loss_MAE_size16 = np.array(loss_MAE_size16)
val_loss_MAE_size16 = np.array(val_loss_MAE_size16)
loss_MAE_size32 = np.array(loss_MAE_size32)
val_loss_MAE_size32 = np.array(val_loss_MAE_size32)
loss_MAE_size64 = np.array(loss_MAE_size64)
val_loss_MAE_size64 = np.array(val_loss_MAE_size64)
loss_MAE_size128 = np.array(loss_MAE_size128)
val_loss_MAE_size128 = np.array(val_loss_MAE_size128)
loss_MAE_size256 = np.array(loss_MAE_size256)
val_loss_MAE_size256 = np.array(val_loss_MAE_size256)

# convert cv_results_ to dataframe
loss_MAE_size16_df = pd.DataFrame(loss_MAE_size16.T)
val_loss_MAE_size16_df = pd.DataFrame(val_loss_MAE_size16.T)
loss_MAE_size32_df = pd.DataFrame(loss_MAE_size32.T)
val_loss_MAE_size32_df = pd.DataFrame(val_loss_MAE_size32.T)
loss_MAE_size64_df = pd.DataFrame(loss_MAE_size64.T)
val_loss_MAE_size64_df = pd.DataFrame(val_loss_MAE_size64.T)
loss_MAE_size128_df = pd.DataFrame(loss_MAE_size128.T)
val_loss_MAE_size128_df = pd.DataFrame(val_loss_MAE_size128.T)
loss_MAE_size256_df = pd.DataFrame(loss_MAE_size256.T)
val_loss_MAE_size256_df = pd.DataFrame(val_loss_MAE_size256.T)

# Save cv_results_ to a CSV file
base_dir = "./20231201/"
os.makedirs(base_dir, exist_ok=True)
cv_results_file = os.path.join(base_dir, 'loss_MAE_size16.csv')
loss_MAE_size16_df.to_csv(cv_results_file, index=False)
cv_results_file = os.path.join(base_dir, 'val_loss_MAE_size16.csv')
val_loss_MAE_size16_df.to_csv(cv_results_file, index=False)
cv_results_file = os.path.join(base_dir, 'loss_MAE_size32.csv')
loss_MAE_size32_df.to_csv(cv_results_file, index=False)
cv_results_file = os.path.join(base_dir, 'val_loss_MAE_size32.csv')
val_loss_MAE_size32_df.to_csv(cv_results_file, index=False)
cv_results_file = os.path.join(base_dir, 'loss_MAE_size64.csv')
loss_MAE_size64_df.to_csv(cv_results_file, index=False)
cv_results_file = os.path.join(base_dir, 'val_loss_MAE_size64.csv')
val_loss_MAE_size64_df.to_csv(cv_results_file, index=False)
cv_results_file = os.path.join(base_dir, 'loss_MAE_size128.csv')
loss_MAE_size128_df.to_csv(cv_results_file, index=False)
cv_results_file = os.path.join(base_dir, 'val_loss_MAE_size128.csv')
val_loss_MAE_size128_df.to_csv(cv_results_file, index=False)
cv_results_file = os.path.join(base_dir, 'loss_MAE_size256.csv')
loss_MAE_size256_df.to_csv(cv_results_file, index=False)
cv_results_file = os.path.join(base_dir, 'val_loss_MAE_size256.csv')
val_loss_MAE_size256_df.to_csv(cv_results_file, index=False)

# Print all the results of CV
print("loss_MAE_size16_mean:\n", np.mean(loss_MAE_size16[:, 199]))
print("val_loss_MAE_size16_mean:\n", np.mean(val_loss_MAE_size16[:, 199]))
print("loss_MAE_size32_mean:\n", np.mean(loss_MAE_size32[:, 199]))
print("val_loss_MAE_size32_mean:\n", np.mean(val_loss_MAE_size32[:, 199]))
print("loss_MAE_size64_mean:\n", np.mean(loss_MAE_size64[:, 199]))
print("val_loss_MAE_size64_mean:\n", np.mean(val_loss_MAE_size64[:, 199]))
print("loss_MAE_size128_mean:\n", np.mean(loss_MAE_size128[:, 199]))
print("val_loss_MAE_size128_mean:\n", np.mean(val_loss_MAE_size128[:, 199]))
print("loss_MAE_size256_mean:\n", np.mean(loss_MAE_size256[:, 199]))
print("val_loss_MAE_size256_mean:\n", np.mean(val_loss_MAE_size256[:, 199]))

# plot the figure of training result
epochs = range(len(loss_MAE_size16.T))
plt.plot(epochs, loss_MAE_size16.T, 'b', label='Training loss')
plt.plot(epochs, val_loss_MAE_size16.T, 'r', label='Validation loss')
plt.title('Training and validation loss')
plt.legend()
plt.figure()

plt.plot(epochs, loss_MAE_size32.T, 'b', label='Training loss')
plt.plot(epochs, val_loss_MAE_size32.T, 'r', label='Validation loss')
plt.title('Training and validation loss')
plt.legend()
plt.figure()

plt.plot(epochs, loss_MAE_size64.T, 'b', label='Training loss')
plt.plot(epochs, val_loss_MAE_size64.T, 'r', label='Validation loss')
plt.title('Training and validation loss')
plt.legend()
plt.figure()

plt.plot(epochs, loss_MAE_size128.T, 'b', label='Training loss')
plt.plot(epochs, val_loss_MAE_size128.T, 'r', label='Validation loss')
plt.title('Training and validation loss')
plt.legend()
plt.figure()

plt.plot(epochs, loss_MAE_size256.T, 'b', label='Training absolute mean error')
plt.plot(epochs, val_loss_MAE_size256.T, 'r', label='Validation absolute mean error')
plt.title('Training and validation absolute mean error')
plt.legend()
plt.figure()

plt.show()
