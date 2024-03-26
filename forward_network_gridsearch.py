# This is a sample Python script.
import time
import datetime
import os
import glob
import sys
import math
from random import shuffle, random, randint

import tensorflow as tf
from keras.layers import *
from keras.models import *
from keras.optimizers import *
from keras.callbacks import *
from keras.wrappers.scikit_learn import KerasRegressor
from keras import backend as K
import numpy as np

from sklearn.model_selection import KFold
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import train_test_split
from sklearn.metrics import make_scorer
from keras.utils import to_categorical
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd


# preprocess of the data
count = 0
h, l, t, dd, ddd = 2.5, 2.0, 0.8, 1.0, 1.0

columns = ["L", "w", "w^3", "alpha", "num_vert", "num_hori", "s1", "s2", "s3", "s4", "s5", "s6", "s7", "s8", "s9", "s10", "s11", "s12", "s13", "s14", "s15", "s16", "s17", "s18", "s19", "s20", "s21", "s22", "s23", "s24", "s25", "s26", "s27", "s28", "s29", "s30", "s31"]
filename = 'processed-data-of-1.6-2.2.xlsx'
data1_1 = pd.read_excel(filename, sheet_name=1, skiprows=None, names=columns)
data1_2 = pd.read_excel(filename, sheet_name=2, skiprows=None, names=columns)
data1_3 = pd.read_excel(filename, sheet_name=3, skiprows=None, names=columns)
data1_4 = pd.read_excel(filename, sheet_name=4, skiprows=None, names=columns)
data1_5 = pd.read_excel(filename, sheet_name=5, skiprows=None, names=columns)
data1 = pd.concat([data1_1, data1_2, data1_3, data1_4, data1_5], axis=0)
data1.pop('s1')
data2 = data1.to_numpy(dtype=np.float32)
data_X = data2[:, :4]
data_Y = data2[:, 6:]

scaler1 = StandardScaler()
scaler1.fit(data_X)
data_X_standarized = scaler1.transform(data_X)

scaler2 = StandardScaler()
scaler2.fit(data_Y)
data_Y_standarized = scaler2.transform(data_Y)

tmp_array = []
for i in range(data2.shape[0]):
    tmp = data_X_standarized[i, :].tolist()
    tmp += to_categorical(data2[i, 4] - 1, 4).tolist()
    tmp += to_categorical(data2[i, 5] - 1, 4).tolist()
    tmp_array += [tmp]
data_X_one_hot = np.array(tmp_array, dtype=np.float32)
data3 = np.append(data_X_one_hot, data_Y_standarized, axis=-1)
data3_augmented = data3

nn = 5
for i in range(1654):
    for j in range(4):
        for k in range(4):
            for m in range(nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1-tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1-(tmp1+tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1-(tmp1+tmp2+tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
# 第二种波形
for i in range(1654, 1845):
    for j in range(4):
        for k in range(4):
            for m in range(9*nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1-tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1-(tmp1+tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1-(tmp1+tmp2+tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
# 第三种波形
for i in range(1845, 1967):
    for j in range(4):
        for k in range(4):
            for m in range(14*nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1-tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1-(tmp1+tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1-(tmp1+tmp2+tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
# 第四种波形
for i in range(1967, 2054):
    for j in range(4):
        for k in range(4):
            for m in range(19*nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1-tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1-(tmp1+tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1-(tmp1+tmp2+tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]

for i in range(2054, 2104):
    for j in range(4):
        for k in range(4):
            for m in range(33*nn):
                if (np.argmax(data3_augmented[i, 4:8]) == j) & (np.argmax(data3_augmented[i, 8:12]) == k):
                    data3_augmented = np.append(data3_augmented, data3[i, :][np.newaxis, :], axis=0)
                    tmp1 = np.random.uniform(low=0.9, high=1.0)
                    tmp1_min = np.min([tmp1, 1-tmp1])
                    tmp2 = np.random.uniform(low=0.0, high=tmp1_min)
                    tmp2_min = np.min([tmp2, 1-(tmp1+tmp2)])
                    tmp3 = np.random.uniform(low=0.0, high=tmp2_min)
                    tmp4 = 1-(tmp1+tmp2+tmp3)
                    tmp = [tmp2, tmp3, tmp4]
                    np.random.shuffle(tmp)
                    tmp5 = np.random.uniform(low=0.9, high=1.0)
                    tmp5_min = np.min([tmp5, 1 - tmp5])
                    tmp6 = np.random.uniform(low=0.0, high=tmp5_min)
                    tmp6_min = np.min([tmp6, 1 - (tmp5 + tmp6)])
                    tmp7 = np.random.uniform(low=0.0, high=tmp6_min)
                    tmp8 = 1 - (tmp5 + tmp6 + tmp7)
                    tmp9 = [tmp6, tmp7, tmp8]
                    np.random.shuffle(tmp9)
                    if (j == 0) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 0) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 5] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 1) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 1) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 6] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 2) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 2) & (k == 3):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 7] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]
                    elif (j == 3) & (k == 0):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 9] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 1):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 10] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    elif (j == 3) & (k == 2):
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 11] = tmp9[2]
                    else:
                        data3_augmented[-1, 4 + j] = tmp1
                        data3_augmented[-1, 4] = tmp[0]
                        data3_augmented[-1, 5] = tmp[1]
                        data3_augmented[-1, 6] = tmp[2]
                        data3_augmented[-1, 8 + k] = tmp5
                        data3_augmented[-1, 8] = tmp9[0]
                        data3_augmented[-1, 9] = tmp9[1]
                        data3_augmented[-1, 10] = tmp9[2]

print(data3_augmented.shape)

np.random.shuffle(data3_augmented)
inputs = data3_augmented[:, :12]
targets = data3_augmented[:, 12:42]

# save augmented data to designated file
# def createwb(wbname):
#     wb=openpyxl.Workbook()
#     wb.save(filename=wbname)
#
#
# def data_toExcel(data, wbname):
#     wb = openpyxl.load_workbook(wbname)
#     worksheet1 = wb.create_sheet(index=0, title='sheet1')
#     for i in range(len(data)):
#         for j in range(len(data[i, :])):
#             worksheet1.cell(row=i+1, column=j+1).value = data[i, j]
#     wb.save(wbname)
#     wb.close()
#
#
# augmented_data_file = 'l2-augmented-1108.xlsx'
# createwb(augmented_data_file)
# data_toExcel(data3_augmented, augmented_data_file)


# Define loss function
def loss_function(y_true, y_pred):
    return K.mean(K.square(y_true - y_pred))


def loss_MAE(y_true, y_pred):
    return float(K.mean(K.abs(y_true - y_pred)))


# Train forward regressor
# def forward_model():
#     inp = Input(shape=(11,), name='surrogator_input')
#
#     x = Dense(128, activation='relu')(inp)
#     x = BatchNormalization()(x)
#
#     x = Dense(256, activation='relu')(x)
#     x = BatchNormalization()(x)
#
#     out = Dense(16, activation=None)(x)
#
#     model = Model(inputs=inp, outputs=out, name='surrogator_points_model')
#
#     return model

# Function to create the forward network
def forward_network(units_layer1=256, units_layer2=900, optimizer='adam'):
    inp = Input(shape=(12,), name='forward_input')
    x = inp
    x = Dense(units_layer1)(x)
    x = BatchNormalization()(x)
    x = Activation('relu')(x)
    x = Dense(units_layer2)(x)
    x = BatchNormalization()(x)
    x = Activation('relu')(x)
    out = Dense(30, activation=None)(x)
    model = Model(inputs=inp, outputs=out)

    if optimizer == 'sgd':
        optimizer = SGD(lr=0.001, momentum=0.9, nesterov=True)
    elif optimizer == 'adam':
        optimizer = Adam(learning_rate=0.001)
    elif optimizer == 'rmsprop':
        optimizer = RMSprop(learning_rate=0.001)

    model.compile(loss=loss_function, optimizer=optimizer, metrics=[loss_MAE])
    return model


# Define the hyperparameter search space
param_grid = {
    'units_layer1': [200, 256, 300],
    'units_layer2': [800, 900, 1024],
    'optimizer': ['adam', 'rmsprop'],
}

# Create the forward_model
forward_model = KerasRegressor(build_fn=forward_network, epochs=200, batch_size=64, verbose=2)

# Create a scorer for the RandomizedSearchCV based on custom loss_MAE
scorer = make_scorer(loss_MAE, greater_is_better=False)

# Split your data into train and test sets
inputs_train, inputs_test, targets_train, targets_test = train_test_split(inputs, targets, test_size=0.2, random_state=42)

# Define a learning rate scheduler callback
lr_scheduler = ReduceLROnPlateau(monitor='loss', factor=0.1, patience=10, min_lr=0.000001, verbose=1)

# Create the RandomizedSearchCV object with the callbacks
random_search = GridSearchCV(forward_model, param_grid, cv=5, scoring=scorer, verbose=2)

print('------------------------------------------------------------------------')
print(f'Training ...')

# Fit the RandomizedSearchCV object to the data
random_search.fit(inputs_train, targets_train, callbacks=[lr_scheduler])

# Print the best hyperparameters
print("Best Hyperparameters: ", random_search.best_params_)
best_hyperparameters = random_search.best_params_
print("all the result of CV: ", random_search.cv_results_)

# Create a new model with the best hyperparameters
best_forward_model = forward_network(units_layer1=best_hyperparameters['units_layer1'], units_layer2=best_hyperparameters['units_layer2'], optimizer=best_hyperparameters['optimizer'])

# Train again the model to obtain the history
train_result = best_forward_model.fit(inputs_train, targets_train, epochs=200, batch_size=32, callbacks=[lr_scheduler], verbose=2, validation_split=0.2)

# save the model of best hyperparameters
base_dir = "./20231130/"
os.makedirs(base_dir, exist_ok=True)
model_loc = os.path.join(base_dir, 'forward_model.keras')
best_forward_model.save(model_loc)

# convert cv_results_ to dataframe
cv_results_df = pd.DataFrame(random_search.cv_results_)

# Save cv_results_ to a CSV file
cv_results_file = os.path.join(base_dir, 'cv_results.csv')
cv_results_df.to_csv(cv_results_file, index=False)

# Print all the results of CV
print("All the results of CV:\n", cv_results_df)

# plot the figure of training result
loss_MAE = train_result.history['loss_MAE']
val_loss_MAE = train_result.history['val_loss_MAE']
loss = train_result.history['loss']
val_loss = train_result.history['val_loss']
epochs = range(len(loss_MAE))

plt.plot(epochs, loss_MAE, 'b', label='Training absolute mean error')
plt.plot(epochs, val_loss_MAE, 'r', label='Validation absolute mean error')
plt.title('Training and validation absolute mean error')
plt.legend()
plt.figure()

plt.plot(epochs, loss, 'b', label='Training loss')
plt.plot(epochs, val_loss, 'r', label='Validation loss')
plt.title('Training and validation loss')
plt.legend()
plt.figure()

# visualize the predicted curve vs ground truth
# test_target_curve_X = data3[1650:1654, :12]
# test_target_curve_Y = data3[1650:1654, 12:]
# test_target_curve_X = np.concatenate((data3[3, :11][np.newaxis, :], data3[7, :11][np.newaxis, :], data3[11, :11][np.newaxis, :], data3[15, :11][np.newaxis, :]), axis=0)
# test_target_curve_Y = np.concatenate((data3[3, 11:][np.newaxis, :], data3[7, 11:][np.newaxis, :], data3[11, 11:][np.newaxis, :], data3[15, 11:][np.newaxis, :]), axis=0)
best_forward_model.load_weights(model_loc, by_name=False)
best_forward_model.trainable = False
points_before_inverse_transform = best_forward_model(inputs_test, training=False)
predicted_points = scaler2.inverse_transform(points_before_inverse_transform)
ground_truth = scaler2.inverse_transform(targets_test)
error_ratio_array = []
for i in range(ground_truth.shape[0]):
    predicted_error = K.sum(K.abs(predicted_points[i, :] - ground_truth[i, :]))
    error_ratio = tf.cast(predicted_error, dtype=tf.float32)/(K.sum(ground_truth[i, :]))
    error_ratio_array.append(error_ratio)
mean_error_ratio = tf.reduce_mean(error_ratio_array)
print(len(error_ratio_array))
print('mean error ratio: ', mean_error_ratio.numpy())
# four_one = plt.subplot(2, 2, 1)
# four_two = plt.subplot(2, 2, 2)
# four_three = plt.subplot(2, 2, 3)
# four_four = plt.subplot(2, 2, 4)
# four_one.scatter(range(len(ground_truth[0, :])), ground_truth[0, :],  c='red', label='Ground truth')
# four_one.scatter(range(len(predicted_points[0, :])), predicted_points[0, :],  c='blue', label='Predicted points')
# four_two.scatter(range(len(ground_truth[1, :])), ground_truth[1, :],  c='red', label='Ground truth')
# four_two.scatter(range(len(predicted_points[1, :])), predicted_points[1, :],  c='blue', label='Predicted points')
# four_three.scatter(range(len(ground_truth[2, :])), ground_truth[2, :],  c='red', label='Ground truth')
# four_three.scatter(range(len(predicted_points[2, :])), predicted_points[2, :],  c='blue', label='Predicted points')
# four_four.scatter(range(len(ground_truth[3, :])), ground_truth[3, :],  c='red', label='Ground truth')
# four_four.scatter(range(len(predicted_points[3, :])), predicted_points[3, :],  c='blue', label='Predicted points')
# plt.title('Ground truth and prediction')
# plt.legend()
plt.show()

# print(loss_MAE, val_loss_MAE)
